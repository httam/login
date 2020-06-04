import tkinter
import os
import pandas as pd
import openpyxl
from openpyxl import Workbook

connected = False
app = True
quit = False

def error():
    global app, var, quit
    error.has_been_called = True
    var = input("souhaitez-vous quitter l'application? ")
    if var == 'oui':
        print(f'merci et à bientot')
        app = False
        quit = True
error.has_been_called = False

def disconnect():
    global app, var, quit
    var = input('souhaitez-vous vous deconnecter? ')
    if var == 'oui':
        print(f'au revoir {user_row[0]} et à bientot! ')
        app = False
        error()

        
if len(os.listdir()) == 1:
    wb = Workbook()
    ws = wb.active
    ws.append(["first_name", "last_name", "email", "password"])
else:
    wb = openpyxl.load_workbook("data.xlsx")
    ws = wb.active

while quit == False:
    while app == True:
        sign_up_sign_in = input('inscription/connexion: ')
        if sign_up_sign_in == "inscription":
            first_name = input('entrez votre prénom: ')
            last_name = input('entrez votre nom: ')
            email = input('entrez votre adresse email: ')
            for row in ws.iter_rows(values_only=True):
                if row[2] == email:
                    print('erreur: adresse email déja utilisée')
                    error()
                    if error.has_been_called == True:
                        break
            if error.has_been_called == True:
                break

            password = input("selectionnez un mot de passe: ")
            while len(password) < 8:
                print('mot de passe trop court')
                password = input("selectionnez un mot de passe: ")
            verification_password = input(
                "veuillez re-selectionner votre mot de passe : ")
            while password != verification_password:
                print('erreur: mots de passe différents')
                verification_password = input(
                    "veuillez re-selectionner votre mot de passe : ")
            ws.append([first_name, last_name, email, password])
            print('inscription terminée!')
            wb.save('data.xlsx')
        if sign_up_sign_in == "connexion":
            email = input('selectionnez votre adresse email: ')
            password = input('entrez votre mot de passe: ')
            for row in ws.iter_rows(values_only=True):
                if row[2] == email:
                    user_row = row
                    break
            if "user_row" not in globals():
                print("erreur: adresse email inconnue")
                error()

                
            else:
                if user_row[-1] != password:
                    print("adresse email ou mot de passe incorrect")
                    error()                

                    
                else:
                    print('connexion réussie!')
                    print(f"bienvenue {user_row[0]}")
                    disconnect()
                    
        